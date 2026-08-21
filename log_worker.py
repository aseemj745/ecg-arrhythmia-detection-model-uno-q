"""
Background Excel logger.

The GUI launches this as a SEPARATE process and pipes one JSON object per
line into its stdin. Each line is a detected arrhythmia episode; this script
appends it to an .xlsx workbook and saves.

    python log_worker.py --excel artifacts/reports/detections.xlsx

Run standalone to test:
    echo {"condition":"PVC","start_s":1.0,"end_s":1.0,"n_beats":1} | python log_worker.py

A separate process rather than a function call, so a slow or failed disk write
can't stall the plot redraw loop, and so the workbook stays valid on disk even
if the GUI is killed mid-demo.
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = [
    ("logged_at", 19),
    ("session", 22),
    ("source", 26),
    ("data_origin", 16),
    ("condition", 11),
    ("start_s", 10),
    ("end_s", 10),
    ("duration_s", 11),
    ("n_beats", 9),
    ("mean_hr_bpm", 13),
    ("confidence", 11),
    ("flag", 16),
    ("annotated_truth", 16),
]

FILLS = {
    "PVC": "FFF2CC",
    "LBBB": "FCE4D6",
    "RBBB": "E2EFDA",
    "AFIB": "F8CBAD",
}


def open_book(path: Path):
    """Load the workbook, or create it with a styled header row."""
    if path.exists():
        try:
            wb = load_workbook(path)
            return wb, wb.active
        except Exception:
            # Corrupt or half-written file: move it aside rather than lose
            # the new session's detections.
            path.rename(path.with_suffix(".corrupt.xlsx"))

    wb = Workbook()
    ws = wb.active
    ws.title = "detections"
    ws.append([h for h, _ in HEADERS])
    for i, (h, w) in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=i)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="4472C4")
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    return wb, ws


def append_event(ws, ev):
    row = [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        ev.get("session", ""),
        ev.get("source", ""),
        ev.get("data_origin", ""),
        ev.get("condition", ""),
        round(float(ev.get("start_s", 0)), 2),
        round(float(ev.get("end_s", 0)), 2),
        round(float(ev.get("end_s", 0)) - float(ev.get("start_s", 0)), 2),
        int(ev.get("n_beats", 0)),
        round(float(ev.get("mean_hr", 0)), 1),
        round(float(ev.get("confidence", 0)), 3),
        ev.get("flag", ""),
        ev.get("annotated_truth", ""),
    ]
    ws.append(row)

    fill = FILLS.get(ev.get("condition", ""))
    if fill:
        ws.cell(row=ws.max_row, column=5).fill = PatternFill("solid",
                                                             fgColor=fill)
    if ev.get("flag"):
        ws.cell(row=ws.max_row, column=12).font = Font(color="C00000",
                                                       italic=True)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", default="artifacts/reports/detections.xlsx")
    ap.add_argument("--flush-every", type=int, default=1,
                    help="save the workbook after this many events")
    args = ap.parse_args()

    path = Path(args.excel)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb, ws = open_book(path)
    wb.save(path)
    print(f"logger ready -> {path}", flush=True)

    pending = 0
    try:
        for line in sys.stdin:
            line = line.strip()
            if not line:
                continue
            try:
                ev = json.loads(line)
            except json.JSONDecodeError:
                print(f"logger: skipped unparseable line: {line[:80]}",
                      flush=True)
                continue

            if ev.get("cmd") == "stop":
                break

            append_event(ws, ev)
            pending += 1
            if pending >= args.flush_every:
                wb.save(path)
                pending = 0
            print(f"logged {ev.get('condition')} @ "
                  f"{float(ev.get('start_s', 0)):.1f}s", flush=True)
    except KeyboardInterrupt:
        pass
    finally:
        wb.save(path)
        print(f"logger stopped, {ws.max_row - 1} rows in {path}", flush=True)


if __name__ == "__main__":
    main()
